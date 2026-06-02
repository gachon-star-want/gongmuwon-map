from __future__ import annotations

import re
import zipfile
from datetime import date, datetime, time
from io import BytesIO
from typing import Any
from xml.etree import ElementTree

import xlrd
from openpyxl import load_workbook
from selectolax.parser import HTMLParser

from public_officer_pipeline import document_guards as guards
from public_officer_pipeline.extractor.rows import RawExpenseFields, build_expense_row
from public_officer_pipeline.models import ParsedExpenseRow


HEADER_ALIASES = {
    "집행일자": "used_date",
    "집행일": "used_date",
    "사용일자": "used_date",
    "사용일": "used_date",
    "승인일": "used_date",
    "승인일자": "used_date",
    "승인일시": "used_date",
    "사용일시": "used_date",
    "집행일시": "used_date",
    "결의일자": "used_date",
    "일시": "used_date",
    "일자": "used_date",
    "일": "used_date",
    "결제일자": "used_date",
    "집행시간": "used_time",
    "사용시간": "used_time",
    "사용시각": "used_time",
    "승인시각": "used_time",
    "승인시간": "used_time",
    "시각": "used_time",
    "시간": "used_time",
    "시": "used_time",
    "부서명": "department_name",
    "부서": "department_name",
    "사용자": "user_text",
    "집행자": "user_text",
    "구분": "user_text",
    "장소": "place_text",
    "사용장소": "place_text",
    "집행장소": "place_text",
    "집행처": "place_text",
    "집행처명": "place_text",
    "사용처": "place_text",
    "가맹점명": "place_text",
    "상호": "place_text",
    "상호명": "place_text",
    "업소명": "place_text",
    "지급처": "place_text",
    "지급처명": "place_text",
    "주소": "address_hint",
    "집행처주소": "address_hint",
    "가맹점주소": "address_hint",
    "집행목적": "purpose",
    "사용목적": "purpose",
    "사용내역": "purpose",
    "카드사용내역": "purpose",
    "사용내용": "purpose",
    "집행내역": "purpose",
    "내용": "purpose",
    "내역": "purpose",
    "대상인원수": "party_size",
    "대상인원": "party_size",
    "집행인원": "party_size",
    "참석자": "party_size",
    "인원": "party_size",
    "인원수": "party_size",
    "결제": "payment_method",
    "금액": "amount",
    "집행금액": "amount",
    "집행액": "amount",
    "지급액": "amount",
    "사용금액": "amount",
    "승인금액": "amount",
    "지출액": "amount",
    "지출금액": "amount",
    "결제방법": "payment_method",
    "결재방법": "payment_method",
    "집행방법": "payment_method",
    "사용방법": "payment_method",
    "방법": "payment_method",
    "사용방법및비고": "payment_method",
    "비목": "expense_category",
    "세목": "expense_category",
    "업무추진비종류": "expense_category",
    "재원": "expense_category",
    "제목": "expense_category",
    "사용액": "amount",
}

DEPARTMENT_RE = re.compile(r"부서명\s*[:：]\s*(?P<department>.+)")
SHEET_MONTH_RE = re.compile(r"(?P<year>20\d{2})\s*(?:년|[./-])\s*(?P<month>\d{1,2})\s*월?")
THOUSAND_UNIT_RE = re.compile(
    r"단위\s*[:：]?\s*천\s*원|금액\s*[（(]\s*천\s*원\s*[)）]|"
    r"집행금액\s*[（(]\s*천\s*원\s*[)）]|지출금액\s*[（(]\s*천\s*원\s*[)）]"
)


def extract_spreadsheet_rows(content: bytes, *, fallback_department: str) -> list[ParsedExpenseRow]:
    return extract_grid_rows(_workbook_rows(content), fallback_department=fallback_department)


def extract_grid_rows(
    grid_rows: list[list[list[str]]],
    *,
    fallback_department: str,
) -> list[ParsedExpenseRow]:
    rows: list[ParsedExpenseRow] = []
    for sheet_rows in grid_rows:
        department = _extract_department(sheet_rows) or fallback_department
        sheet_month = _extract_sheet_month(sheet_rows)
        header_index, mapped_headers = _find_header(sheet_rows)
        if header_index is None:
            continue
        amount_is_thousands = _sheet_amount_is_thousands(sheet_rows, header_index, mapped_headers)
        for raw_row in sheet_rows[header_index + 1 :]:
            parsed = _parse_row(
                raw_row,
                mapped_headers,
                department,
                sheet_month=sheet_month,
                amount_is_thousands=amount_is_thousands,
            )
            if parsed:
                rows.append(parsed)
    return rows


def _workbook_rows(content: bytes) -> list[list[list[str]]]:
    guards.ensure_size_at_most(
        size=len(content),
        max_bytes=guards.MAX_SPREADSHEET_BYTES,
        subject="spreadsheet document",
    )
    if content.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return _xls_workbook_rows(content)
    try:
        guards.preflight_xlsx_zip(content)
    except zipfile.BadZipFile:
        html_rows = _html_table_rows(content)
        if html_rows:
            return html_rows
        raise
    workbook = _load_xlsx_workbook(content)
    try:
        worksheets = workbook.worksheets[: guards.MAX_SPREADSHEET_SHEETS]
        workbook_rows: list[list[list[str]]] = []
        total_cells = 0
        for worksheet in worksheets:
            sheet_rows: list[list[str]] = []
            empty_run = 0
            non_empty_row_count = 0
            for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                row_values = _trim_trailing_empty_cells([_stringify(cell) for cell in row])
                if not any(row_values):
                    empty_run += 1
                    if sheet_rows and empty_run >= guards.MAX_SPREADSHEET_ROWS_PER_SHEET:
                        break
                    continue
                empty_run = 0
                non_empty_row_count += 1
                if non_empty_row_count > guards.MAX_SPREADSHEET_ROWS_PER_SHEET:
                    raise guards.DocumentProcessingLimitError(
                        f"spreadsheet sheet {worksheet.title!r} rows count exceeds limit of "
                        f"{guards.MAX_SPREADSHEET_ROWS_PER_SHEET}"
                    )
                total_cells += _checked_row_width(
                    sheet_name=worksheet.title,
                    row_index=non_empty_row_count,
                    width=len(row_values),
                    current_total_cells=total_cells,
                )
                sheet_rows.append(row_values)
            workbook_rows.append(sheet_rows)
        return workbook_rows
    finally:
        workbook.close()


def _trim_trailing_empty_cells(row_values: list[str]) -> list[str]:
    last_non_empty = -1
    non_empty_count = 0
    for index, value in enumerate(row_values):
        if value:
            last_non_empty = index
            non_empty_count += 1
    trimmed = row_values[: last_non_empty + 1]
    if (
        len(trimmed) > guards.MAX_SPREADSHEET_COLUMNS_PER_SHEET
        and non_empty_count <= guards.MAX_SPREADSHEET_COLUMNS_PER_SHEET
    ):
        return trimmed[: guards.MAX_SPREADSHEET_COLUMNS_PER_SHEET]
    return trimmed


def _load_xlsx_workbook(content: bytes):
    repaired = content
    for _ in range(3):
        try:
            return load_workbook(BytesIO(repaired), data_only=True, read_only=True)
        except TypeError as exc:
            updated = _repair_xlsx_for_openpyxl_type_error(repaired, str(exc))
            if updated == repaired:
                raise
            repaired = updated
    return load_workbook(BytesIO(repaired), data_only=True, read_only=True)


def _repair_xlsx_for_openpyxl_type_error(content: bytes, message: str) -> bytes:
    if "_NamedCellStyle" in message and "name should be" in message:
        return _repair_xlsx_missing_cell_style_names(content)
    if "StringProperty" in message and "name should be" in message:
        return _repair_xlsx_missing_custom_property_names(content)
    return content


def _repair_xlsx_missing_cell_style_names(content: bytes) -> bytes:
    with zipfile.ZipFile(BytesIO(content)) as source:
        if "xl/styles.xml" not in source.namelist():
            return content
        repaired_styles = _repair_styles_xml_missing_cell_style_names(source.read("xl/styles.xml"))
        if repaired_styles is None:
            return content
        output = BytesIO()
        with zipfile.ZipFile(output, "w") as target:
            for item in source.infolist():
                data = repaired_styles if item.filename == "xl/styles.xml" else source.read(item.filename)
                target.writestr(item, data)
    return output.getvalue()


def _repair_xlsx_missing_custom_property_names(content: bytes) -> bytes:
    with zipfile.ZipFile(BytesIO(content)) as source:
        if "docProps/custom.xml" not in source.namelist():
            return content
        repaired_props = _repair_custom_xml_missing_property_names(
            source.read("docProps/custom.xml")
        )
        if repaired_props is None:
            return content
        output = BytesIO()
        with zipfile.ZipFile(output, "w") as target:
            for item in source.infolist():
                data = repaired_props if item.filename == "docProps/custom.xml" else source.read(item.filename)
                target.writestr(item, data)
    return output.getvalue()


def _repair_custom_xml_missing_property_names(custom_xml: bytes) -> bytes | None:
    try:
        root = ElementTree.fromstring(custom_xml)
    except ElementTree.ParseError:
        return None

    namespace = root.tag[1:].split("}", 1)[0] if root.tag.startswith("{") else ""
    property_tag = f"{{{namespace}}}property" if namespace else "property"
    existing_names = {element.attrib["name"] for element in root.iter(property_tag) if element.attrib.get("name")}
    repaired = False
    for index, element in enumerate(root.iter(property_tag), start=1):
        if element.attrib.get("name"):
            continue
        name = _unique_cell_style_name(f"Recovered Custom Property {index}", existing_names)
        element.set("name", name)
        existing_names.add(name)
        repaired = True
    if not repaired:
        return None
    return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def _repair_styles_xml_missing_cell_style_names(styles_xml: bytes) -> bytes | None:
    try:
        root = ElementTree.fromstring(styles_xml)
    except ElementTree.ParseError:
        return None

    namespace = root.tag[1:].split("}", 1)[0] if root.tag.startswith("{") else ""
    cell_style_tag = f"{{{namespace}}}cellStyle" if namespace else "cellStyle"
    existing_names = {element.attrib["name"] for element in root.iter(cell_style_tag) if element.attrib.get("name")}
    repaired = False
    for index, element in enumerate(root.iter(cell_style_tag), start=1):
        if element.attrib.get("name"):
            continue
        base_name = _builtin_cell_style_name(element.attrib.get("builtinId")) or f"Recovered Style {index}"
        name = _unique_cell_style_name(base_name, existing_names)
        element.set("name", name)
        existing_names.add(name)
        repaired = True
    if not repaired:
        return None
    return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def _builtin_cell_style_name(builtin_id: str | None) -> str | None:
    return {
        "0": "Normal",
        "3": "Comma",
        "4": "Currency",
        "5": "Percent",
        "6": "Comma [0]",
        "7": "Currency [0]",
    }.get(builtin_id or "")


def _unique_cell_style_name(base_name: str, existing_names: set[str]) -> str:
    if base_name not in existing_names:
        return base_name
    suffix = 2
    while f"{base_name} {suffix}" in existing_names:
        suffix += 1
    return f"{base_name} {suffix}"


def _xls_workbook_rows(content: bytes) -> list[list[list[str]]]:
    workbook = xlrd.open_workbook(file_contents=content)
    worksheets: list[list[list[str]]] = []
    total_cells = 0
    for worksheet in workbook.sheets()[: guards.MAX_SPREADSHEET_SHEETS]:
        if worksheet.nrows > guards.MAX_SPREADSHEET_ROWS_PER_SHEET:
            raise guards.DocumentProcessingLimitError(
                f"spreadsheet sheet {worksheet.name!r} has {worksheet.nrows} rows, "
                f"exceeding limit of {guards.MAX_SPREADSHEET_ROWS_PER_SHEET}"
            )
        rows: list[list[str]] = []
        non_empty_row_count = 0
        for row_index in range(worksheet.nrows):
            row_values = _trim_trailing_empty_cells(
                [
                    _stringify(
                        _xls_cell_value(worksheet.cell(row_index, column_index), workbook.datemode)
                    )
                    for column_index in range(worksheet.ncols)
                ]
            )
            if not any(row_values):
                continue
            non_empty_row_count += 1
            total_cells += _checked_row_width(
                sheet_name=worksheet.name,
                row_index=non_empty_row_count,
                width=len(row_values),
                current_total_cells=total_cells,
            )
            rows.append(row_values)
        worksheets.append(rows)
    return worksheets


def _html_table_rows(content: bytes) -> list[list[list[str]]]:
    text = _decode_htmlish_content(content)
    if "<table" not in text.lower():
        return []

    tree = HTMLParser(text)
    tables: list[list[list[str]]] = []
    total_cells = 0
    for table_index, table in enumerate(tree.css("table"), start=1):
        if table_index > guards.MAX_SPREADSHEET_SHEETS:
            break
        sheet_name = f"html table {table_index}"
        rows: list[list[str]] = []
        for tr in table.css("tr"):
            row_values = _trim_trailing_empty_cells(
                [_clean(cell.text(separator=" ", strip=True)) for cell in tr.css("th,td")]
            )
            if not any(row_values):
                continue
            row_index = len(rows) + 1
            if row_index > guards.MAX_SPREADSHEET_ROWS_PER_SHEET:
                raise guards.DocumentProcessingLimitError(
                    f"spreadsheet sheet {sheet_name!r} rows count exceeds limit of "
                    f"{guards.MAX_SPREADSHEET_ROWS_PER_SHEET}"
                )
            total_cells += _checked_row_width(
                sheet_name=sheet_name,
                row_index=row_index,
                width=len(row_values),
                current_total_cells=total_cells,
            )
            rows.append(row_values)
        if rows:
            tables.append(rows)
    return tables


def _decode_htmlish_content(content: bytes) -> str:
    for encoding in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _ensure_declared_sheet_bounds(
    *,
    sheet_name: str,
    rows: int,
    columns: int,
    current_total_cells: int,
) -> int:
    if rows > guards.MAX_SPREADSHEET_ROWS_PER_SHEET:
        raise guards.DocumentProcessingLimitError(
            f"spreadsheet sheet {sheet_name!r} has {rows} rows, "
            f"exceeding limit of {guards.MAX_SPREADSHEET_ROWS_PER_SHEET}"
        )
    if columns > guards.MAX_SPREADSHEET_COLUMNS_PER_SHEET:
        raise guards.DocumentProcessingLimitError(
            f"spreadsheet sheet {sheet_name!r} has {columns} columns, "
            f"exceeding limit of {guards.MAX_SPREADSHEET_COLUMNS_PER_SHEET}"
        )
    total_cells = current_total_cells + (rows * columns)
    if total_cells > guards.MAX_SPREADSHEET_CELLS_TOTAL:
        raise guards.DocumentProcessingLimitError(
            f"spreadsheet workbook has {total_cells} declared cells, "
            f"exceeding limit of {guards.MAX_SPREADSHEET_CELLS_TOTAL}"
        )
    return total_cells


def _checked_row_width(
    *,
    sheet_name: str,
    row_index: int,
    width: int,
    current_total_cells: int,
) -> int:
    if row_index > guards.MAX_SPREADSHEET_ROWS_PER_SHEET:
        raise guards.DocumentProcessingLimitError(
            f"spreadsheet sheet {sheet_name!r} rows count exceeds limit of "
            f"{guards.MAX_SPREADSHEET_ROWS_PER_SHEET}"
        )
    if width > guards.MAX_SPREADSHEET_COLUMNS_PER_SHEET:
        raise guards.DocumentProcessingLimitError(
            f"spreadsheet sheet {sheet_name!r} row {row_index} has {width} columns, "
            f"exceeding limit of {guards.MAX_SPREADSHEET_COLUMNS_PER_SHEET}"
        )
    if current_total_cells + width > guards.MAX_SPREADSHEET_CELLS_TOTAL:
        raise guards.DocumentProcessingLimitError(
            f"spreadsheet workbook cells count exceeds limit of "
            f"{guards.MAX_SPREADSHEET_CELLS_TOTAL}"
        )
    return width


def _xls_cell_value(cell: xlrd.sheet.Cell, datemode: int) -> Any:
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate.xldate_as_datetime(cell.value, datemode)
    return cell.value


def _extract_department(rows: list[list[str]]) -> str | None:
    for row in rows[:10]:
        text = " ".join(cell for cell in row if cell)
        match = DEPARTMENT_RE.search(text)
        if match:
            return _clean(match.group("department"))
    return None


def _extract_sheet_month(rows: list[list[str]]) -> tuple[int, int] | None:
    for row in rows[:10]:
        text = " ".join(cell for cell in row if cell)
        match = SHEET_MONTH_RE.search(text)
        if not match:
            continue
        try:
            return int(match.group("year")), int(match.group("month"))
        except ValueError:
            continue
    return None


def _find_header(rows: list[list[str]]) -> tuple[int | None, list[str | None]]:
    for index, row in enumerate(rows[:20]):
        mapped = _disambiguate_headers(row, [_map_header(cell) for cell in row])
        if index + 1 < len(rows):
            width = max(len(row), len(rows[index + 1]))
            overlaid = [
                (rows[index + 1][column] if column < len(rows[index + 1]) else "")
                or (row[column] if column < len(row) else "")
                for column in range(width)
            ]
            overlaid_mapped = _disambiguate_headers(overlaid, [_map_header(cell) for cell in overlaid])
            if (
                any(mapped)
                and _has_required_headers(overlaid_mapped)
                and _header_score(overlaid_mapped) >= _header_score(mapped)
            ):
                return index + 1, overlaid_mapped
        if _has_required_headers(mapped):
            return index, mapped
    return None, []


def _has_required_headers(mapped: list[str | None]) -> bool:
    has_place_hint = "place_text" in mapped or "purpose" in mapped or "expense_category" in mapped
    return "used_date" in mapped and has_place_hint and "amount" in mapped


def _header_score(mapped: list[str | None]) -> int:
    return sum(1 for header in mapped if header)


def _map_header(header: str) -> str | None:
    compact = re.sub(r"\s+", "", header)
    normalized = re.sub(r"[（(][^)）]+[)）]", "", compact)
    if normalized.startswith(("집행목적", "사용목적", "집행내역")):
        return "purpose"
    if normalized.startswith(("사용장소", "집행장소")):
        return "place_text"
    return HEADER_ALIASES.get(normalized) or HEADER_ALIASES.get(compact)


def _disambiguate_headers(raw_headers: list[str], mapped_headers: list[str | None]) -> list[str | None]:
    mapped = list(mapped_headers)
    compact = [re.sub(r"\s+", "", header) for header in raw_headers]
    if "사용자" in compact and compact and compact[0] == "구분":
        mapped[0] = "expense_category"
    execution_type_indexes = [index for index, header in enumerate(compact) if header == "집행유형"]
    if len(execution_type_indexes) >= 2:
        mapped[execution_type_indexes[0]] = "purpose"
        mapped[execution_type_indexes[-1]] = "payment_method"
    elif len(execution_type_indexes) == 1:
        mapped[execution_type_indexes[0]] = "purpose"
    if "지급처1" in compact and "지급처2" in compact:
        mapped[compact.index("지급처1")] = "address_hint"
        mapped[compact.index("지급처2")] = "place_text"
    return mapped


def _parse_row(
    raw_row: list[str],
    mapped_headers: list[str | None],
    department: str,
    *,
    sheet_month: tuple[int, int] | None = None,
    amount_is_thousands: bool = False,
) -> ParsedExpenseRow | None:
    item = {
        mapped_headers[index]: _clean(value)
        for index, value in enumerate(raw_row[: len(mapped_headers)])
        if mapped_headers[index] and _clean(value)
    }
    place_text = item.get("place_text") or item.get("purpose") or item.get("expense_category")
    if not item.get("used_date") or not place_text or not item.get("amount"):
        return None

    department_name = item.get("department_name") or department
    user_text = item.get("user_text") or department_name
    amount_text = item.get("amount")
    date_text = _row_date_text(raw_row, mapped_headers, item["used_date"], sheet_month=sheet_month)
    return build_expense_row(
        RawExpenseFields(
            department_name=department_name,
            date_text=date_text,
            time_text=item.get("used_time"),
            place_name=place_text,
            address=item.get("address_hint"),
            purpose=item.get("purpose") or None,
            amount=amount_text,
            amount_is_thousands=_should_scale_amount_as_thousands(amount_text, amount_is_thousands),
            party_size=item.get("party_size"),
            user_text=user_text,
            payment_method=item.get("payment_method") or None,
            expense_category=item.get("expense_category") or None,
            raw_values=raw_row,
        ),
        fallback_department=department,
        address_separator=" (",
    )


def _sheet_amount_is_thousands(
    rows: list[list[str]],
    header_index: int,
    mapped_headers: list[str | None],
) -> bool:
    header_row = rows[header_index] if header_index < len(rows) else []
    for index, mapped in enumerate(mapped_headers):
        if mapped == "amount" and index < len(header_row):
            if THOUSAND_UNIT_RE.search(re.sub(r"\s+", "", header_row[index])):
                return True
    for row in rows[: max(header_index + 1, 10)]:
        text = re.sub(r"\s+", "", " ".join(cell for cell in row if cell))
        if THOUSAND_UNIT_RE.search(text):
            return True
    return False


def _should_scale_amount_as_thousands(value: str | None, amount_is_thousands: bool) -> bool:
    if not amount_is_thousands or value is None:
        return False
    numeric = re.sub(r"[^\d]", "", str(value))
    if not numeric:
        return False
    return int(numeric) < 10_000


def _row_date_text(
    raw_row: list[str],
    mapped_headers: list[str | None],
    used_date: str,
    *,
    sheet_month: tuple[int, int] | None,
) -> str:
    if sheet_month is None or not re.fullmatch(r"\d{1,2}", used_date):
        return used_date
    try:
        date_index = mapped_headers.index("used_date")
    except ValueError:
        return used_date

    year, sheet_month_value = sheet_month
    next_value = _clean(raw_row[date_index + 1]) if date_index + 1 < len(raw_row) else ""
    if re.fullmatch(r"\d{1,2}", next_value):
        first = int(used_date)
        day = int(next_value)
        month = first if 1 <= first <= 12 else sheet_month_value
        if first == sheet_month_value:
            month = sheet_month_value
    else:
        month = sheet_month_value
        day = int(used_date)
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return used_date


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return str(value)


def _clean(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()
