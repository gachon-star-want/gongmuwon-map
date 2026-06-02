import re
import zipfile
from io import BytesIO

import pytest
from openpyxl import Workbook

from public_officer_pipeline import document_guards as guards
from public_officer_pipeline.extractor import spreadsheet as spreadsheet_module
from public_officer_pipeline.extractor import extract_spreadsheet_rows


def _workbook_bytes(workbook: Workbook) -> bytes:
    content = BytesIO()
    workbook.save(content)
    return content.getvalue()


def _zip_bytes(entries: dict[str, bytes]) -> bytes:
    content = BytesIO()
    with zipfile.ZipFile(content, "w") as archive:
        for name, value in entries.items():
            archive.writestr(name, value)
    return content.getvalue()


def _workbook_bytes_without_normal_style_name(workbook: Workbook) -> bytes:
    original = _workbook_bytes(workbook)
    output = BytesIO()
    with zipfile.ZipFile(BytesIO(original)) as source, zipfile.ZipFile(output, "w") as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "xl/styles.xml":
                data, count = re.subn(
                    rb'(<[^:>]*:?cellStyle\b)([^>]*?)\sname="Normal"([^>]*/?>)',
                    rb"\1\2\3",
                    data,
                    count=1,
                )
                assert count == 1
            target.writestr(item, data)
    return output.getvalue()


def test_extracts_gangnam_xlsx_rows() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "공개"
    worksheet.append(["2026년 4월 업무추진비 집행내역"])
    worksheet.append(["□ 부서명 : 지방소득세과"])
    worksheet.append([])
    worksheet.append(["연번", "집행일자", "집행시간", "사용자", "장소", "집행 목적", "대상 인원수", "금액", "결제방법", "비목"])
    worksheet.append(
        [
            1,
            "2026-04-30",
            "19:47:37",
            "지방소득세과장",
            "남도사계고운님",
            "세무서 업무 협의",
            6,
            121000,
            "카드결제",
            "시책",
        ]
    )
    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="강남구청")

    assert len(rows) == 1
    assert rows[0].department_name == "지방소득세과"
    assert rows[0].place_text == "남도사계고운님"
    assert rows[0].amount == 121000
    assert rows[0].user_text == "지방소득세과장 6명"


def test_extracts_spreadsheet_amounts_marked_in_thousand_won() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["2025년 기관장 업무추진비 세부집행내역"])
    worksheet.append(["(단위 : 천원)"])
    worksheet.append(["사용일자", "집행목적", "장소", "대상 인원수(명)", "지출금액(천원)"])
    worksheet.append(["2025-07-04", "업무협의", "테스트식당", 4, 160])

    rows = extract_spreadsheet_rows(_workbook_bytes(workbook), fallback_department="테스트기관")

    assert len(rows) == 1
    assert rows[0].amount == 160000


def test_ignores_bare_day_without_sheet_month_context() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["기관장 업무추진비 집행내역"])
    worksheet.append(["일자", "장소", "집행목적", "금액"])
    worksheet.append(["4", "4건", "4건", 316400])

    rows = extract_spreadsheet_rows(_workbook_bytes(workbook), fallback_department="테스트기관")

    assert rows == []


def test_extracts_bare_day_when_sheet_month_context_exists() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["2025년 6월 기관장 업무추진비 집행내역"])
    worksheet.append(["일자", "장소", "집행목적", "금액"])
    worksheet.append(["10", "테스트식당", "업무협의", 70000])

    rows = extract_spreadsheet_rows(_workbook_bytes(workbook), fallback_department="테스트기관")

    assert len(rows) == 1
    assert rows[0].used_at.date().isoformat() == "2025-06-10"


def test_keeps_large_won_amounts_when_sheet_mentions_thousand_unit_summary() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["2025년 기관장 업무추진비"])
    worksheet.append(["(단위 : 천원)"])
    worksheet.append(["사용일자", "집행목적", "장소", "대상 인원수(명)", "지출금액(원)"])
    worksheet.append(["2025-07-04", "업무협의", "테스트식당", 4, 160000])

    rows = extract_spreadsheet_rows(_workbook_bytes(workbook), fallback_department="테스트기관")

    assert len(rows) == 1
    assert rows[0].amount == 160000


def test_extracts_council_cost_xlsx_header_variants() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "부의장"
    worksheet.append(["부의장 활동비 집행내역(4월)"])
    worksheet.append([])
    worksheet.append(["사용자", "일 자", "시간", "집행처 명", "집행내역", "집행처 주소", "집행방법", "금액", "인원", "비고"])
    worksheet.append(
        [
            "이충현",
            "2026-04-01",
            "12:29",
            "당진아구동태찜탕",
            "의정활동설명 및 민원청취",
            "양천구 등촌로 182",
            "법인카드",
            55000,
            4,
            None,
        ]
    )
    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="강서구의회")

    assert len(rows) == 1
    assert rows[0].place_text == "당진아구동태찜탕 (양천구 등촌로 182)"
    assert rows[0].purpose == "의정활동설명 및 민원청취"
    assert rows[0].payment_method == "법인카드"
    assert rows[0].user_text == "이충현 4명"


def test_extracts_council_cost_xlsx_two_row_headers() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "집행내역"
    worksheet.append(["2026년 4월 의장단 업무추진비 집행내역"])
    worksheet.append(
        [
            "연번",
            "비목",
            "집행부서",
            "사용자",
            "집행일시",
            None,
            "집행금액\n(원)",
            "집행장소",
            None,
            "집행내역",
            "대상\n인원수",
            "결제방법",
        ]
    )
    worksheet.append([None, None, None, None, "일자", "시간", None, "상호명", "주소", None, None, None])
    worksheet.append(
        [
            1,
            "의회운영업무추진비",
            "관악구의회",
            "의장",
            "2026.04.01",
            "13:12",
            19000,
            "만리장성",
            "서울특별시 관악구 관악로 146",
            "의정활동 및 의회운영 관련 업무유대 경비",
            2,
            "신용카드",
        ]
    )
    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="관악구의회")

    assert len(rows) == 1
    assert rows[0].place_text == "만리장성 (서울특별시 관악구 관악로 146)"
    assert rows[0].amount == 19000
    assert rows[0].purpose == "의정활동 및 의회운영 관련 업무유대 경비"
    assert rows[0].user_text == "의장 2명"


def test_extracts_council_cost_xlsx_approval_amount_header() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "지출내역"
    worksheet.append(["2026년 3월 의장단 업무추진비 집행내역"])
    worksheet.append([])
    worksheet.append(["연번", "일자", "시간", "승인금액", "장소", "주소", "내역", "대상인원", "결제방법", "사용자"])
    worksheet.append(
        [
            1,
            "2026-03-04",
            "12:27:40",
            78000,
            "보승회관 구로구청점",
            "서울특별시 구로구 구로중앙로 68",
            "지역 현안 의견청취를 위한 간담회",
            6,
            "카드",
            "정대근",
        ]
    )
    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="구로구의회")

    assert len(rows) == 1
    assert rows[0].amount == 78000
    assert rows[0].place_text == "보승회관 구로구청점 (서울특별시 구로구 구로중앙로 68)"
    assert rows[0].user_text == "정대근 6명"


def test_extracts_yangju_council_approval_card_usage_headers() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "의장"
    worksheet.append(["2026년 1분기 의장 업무추진비 사용내역"])
    worksheet.append([])
    worksheet.append(["연번", "부서명", "승인일자", "승인시간", "승인금액", "가맹점명", "카드사용내역", "인원"])
    worksheet.append([1, "의회사무과", "2026-01-05", "13:09", 46800, "이디야커피 양주장흥점", "장흥면 이장협의회 후 현안사항 간담회", 14])
    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="양주시의회")

    assert len(rows) == 1
    assert rows[0].department_name == "의회사무과"
    assert rows[0].place_text == "이디야커피 양주장흥점"
    assert rows[0].purpose == "장흥면 이장협의회 후 현안사항 간담회"
    assert rows[0].amount == 46800
    assert rows[0].user_text == "의회사무과 14명"


def test_extracts_council_cost_xlsx_expense_amount_header() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "의회사무과"
    worksheet.append(["업무추진비 사용내역"])
    worksheet.append(["<2026년 4월/의회사무과장/기관>"])
    worksheet.append([])
    worksheet.append(["일시", "집행목적", "장소(사용처)", "방법", "인원", "지출액"])
    worksheet.append(
        [
            "2026-04-22 15:25:09",
            "업무추진 직원 격려 식사",
            "굽네치킨 여주점",
            "신용카드",
            19,
            182300,
        ]
    )
    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="여주시의회")

    assert len(rows) == 1
    assert rows[0].department_name == "여주시의회"
    assert rows[0].place_text == "굽네치킨 여주점"
    assert rows[0].purpose == "업무추진 직원 격려 식사"
    assert rows[0].payment_method == "신용카드"
    assert rows[0].amount == 182300
    assert rows[0].user_text == "여주시의회 19명"


def test_extracts_daejeon_xlsx_date_with_weekday_suffix() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "시책추진"
    worksheet.append(["업무추진비 사용내역('26.5월)"])
    worksheet.append(
        [
            "사용자",
            "사용일자(일시)",
            "사용장소\n(가맹점명)",
            "사용목적(내역)\n*사용대상 포함",
            "사용금액(원)",
            "대상인원(명)",
            "사용방법",
        ]
    )
    worksheet.append(
        [
            "환경국장",
            "2026. 5. 8.(금) 12:00",
            "서가앤쿡,스타벅스 / 서구 둔산동",
            "환경정책 관련 의견수렴 및 정책반영 등",
            78900,
            4,
            "카드",
        ]
    )
    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="대전시청 환경국")

    assert len(rows) == 1
    assert rows[0].used_at.isoformat() == "2026-05-08T12:00:00"
    assert rows[0].department_name == "대전시청 환경국"
    assert rows[0].place_text == "서가앤쿡,스타벅스 / 서구 둔산동"
    assert rows[0].amount == 78900
    assert rows[0].user_text == "환경국장 4명"
    assert rows[0].payment_method == "카드"


def test_extracts_uiwang_xlsx_usage_detail_merchant_headers() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "국장"
    worksheet.append(["시책추진업무추진비 사용내역(2월)"])
    worksheet.append(["안전환경교통국", None, None, None, None, "(단위 : 원)"])
    worksheet.append(["사용일시", "사 용 내 역", "참석자", "사용처", "사 용 금 액", "비 고"])
    worksheet.append(["합계", None, None, None, 1313400])
    worksheet.append(
        [
            "2026. 2. 4.(수) 20:45",
            "백운호수 생태로 탐방 사업 관련하여 관계자와의 간담회",
            "관계자 등 10명",
            "천장어",
            352400,
            None,
        ]
    )
    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="의왕시청 안전총괄과")

    assert len(rows) == 1
    assert rows[0].used_at.isoformat() == "2026-02-04T20:45:00"
    assert rows[0].department_name == "의왕시청 안전총괄과"
    assert rows[0].place_text == "천장어"
    assert rows[0].purpose == "백운호수 생태로 탐방 사업 관련하여 관계자와의 간담회"
    assert rows[0].amount == 352400
    assert rows[0].user_text == "의왕시청 안전총괄과 10명"


def test_extracts_incheon_junggu_xlsx_payment_date_usage_content_headers() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "업무추진비 (구)"
    worksheet.append(["2026. 4월 업무추진비 사용내역(인천 중구청장)"])
    worksheet.append([])
    worksheet.append(["연번", "결제일자", "시간", "금액(원)", "사용내용", "업소명", "인원수", "업무추진비 종류", "사용자", "결제방법"])
    worksheet.append(
        [
            1,
            "2026-04-01 00:00:00",
            "12:27",
            234000,
            "주민자치 협의회 운영 등 논의에 따른 간담회 업무추진비 지출",
            "신포바다애",
            14,
            "시책",
            "구청장",
            "카드",
        ]
    )
    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="인천광역시 중구청")

    assert len(rows) == 1
    assert rows[0].used_at.isoformat() == "2026-04-01T12:27:00"
    assert rows[0].place_text == "신포바다애"
    assert rows[0].purpose == "주민자치 협의회 운영 등 논의에 따른 간담회 업무추진비 지출"
    assert rows[0].expense_category == "시책"
    assert rows[0].user_text == "구청장 14명"


def test_extracts_incheon_junggu_council_xlsx_content_and_spaced_place_headers() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "의장"
    worksheet.append(["업무추진비 사용내역"])
    worksheet.append([])
    worksheet.append(["연번", "결제일자", "시간", "제  목", "금액(원)", "내  용", "장  소", "인원수", "결제방법"])
    worksheet.append(["", "합계", "", "", 6329000])
    worksheet.append(
        [
            1,
            "2026-01-04 00:00:00",
            "18:57",
            "의회운영업무추진비(의장) 지출",
            86000,
            "제3연륙교 행사 참석후 직원 노고 격려",
            "신흥부대고기",
            6,
            "카드",
        ]
    )
    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="인천광역시 중구의회")

    assert len(rows) == 1
    assert rows[0].used_at.isoformat() == "2026-01-04T18:57:00"
    assert rows[0].place_text == "신흥부대고기"
    assert rows[0].purpose == "제3연륙교 행사 참석후 직원 노고 격려"
    assert rows[0].expense_category == "의회운영업무추진비(의장) 지출"
    assert rows[0].user_text == "인천광역시 중구의회 6명"


def test_extracts_michuhol_xlsx_split_month_day_usage_amount_headers() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "부서운영"
    worksheet.append(["기초생활보장과 업무추진비 집행현황(2026. 5월)"])
    worksheet.append(["▣ 총 1건  126,100원"])
    worksheet.append(["▣ 세부내역", None, None, None, None, None, None, None, "[단위: 원]"])
    worksheet.append([])
    worksheet.append(["세 목", "사용일자", None, "사용처", "사용방법", "사용목적", "사용대상", "인원", "사용액"])
    worksheet.append([])
    worksheet.append(["부서운영", "계", None, None, None, None, None, None, 126100])
    worksheet.append(["", 5, 7, "코차 인터내셔널", "법인카드", "차류 구입", "직원", "39명", 126100])
    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="미추홀구청 기초생활보장과")

    assert len(rows) == 1
    assert rows[0].used_at.isoformat() == "2026-05-07T00:00:00"
    assert rows[0].department_name == "미추홀구청 기초생활보장과"
    assert rows[0].place_text == "코차 인터내셔널"
    assert rows[0].purpose == "차류 구입"
    assert rows[0].amount == 126100
    assert rows[0].user_text == "미추홀구청 기초생활보장과 39명"
    assert rows[0].payment_method == "법인카드"


def test_extracts_council_cost_xlsx_short_day_time_amount_headers() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "의장단"
    worksheet.append(["2026년 4월 의장단 업무추진비 집행내역"])
    worksheet.append([])
    worksheet.append(["번호", "사용자", "일", "시", "장소", "집행목적", "대상인원수", "금액(원)", "결제방법", "비목"])
    worksheet.append(
        [
            1,
            "의장",
            "2026-04-01 00:00:00",
            "1899-12-31 12:10:00",
            "차순옥",
            "결산검사위원과의 간담회",
            12,
            240000,
            "카드결제",
            "의회운영",
        ]
    )
    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="평택시의회")

    assert len(rows) == 1
    assert rows[0].used_at.isoformat() == "2026-04-01T12:10:00"
    assert rows[0].place_text == "차순옥"
    assert rows[0].purpose == "결산검사위원과의 간담회"
    assert rows[0].amount == 240000
    assert rows[0].user_text == "의장 12명"


def test_extracts_council_cost_xlsx_short_merchant_and_party_headers() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "의장"
    worksheet.append(["의회명", "구분", "사용일", "사용일시", "상호", "주소", "집행목적", "사용금액(원)", "인원수", "결제방법"])
    worksheet.append(
        [
            "도봉구의회",
            "의장",
            "2026-01-05",
            "12:22:35",
            "은행골",
            "서울 도봉구 도봉동 635번지",
            "지방의회 운영 방향 등 정책논의 간담회",
            180000,
            9,
            "카드",
        ]
    )
    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="도봉구의회")

    assert len(rows) == 1
    assert rows[0].place_text == "은행골 (서울 도봉구 도봉동 635번지)"
    assert rows[0].amount == 180000
    assert rows[0].user_text == "의장 9명"


def test_extracts_council_cost_xlsx_duplicate_execution_type_headers() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "의장"
    worksheet.append(["의장 업무추진비 집행내역"])
    worksheet.append([])
    worksheet.append(["구분", "사용자", "집행일", "집행유형", "집행구분", "집행대상", "집행인원", "집행액(천원)", "장소", "시간", "집행유형"])
    worksheet.append(
        [
            "의회운영",
            "의장",
            "2026-03-03",
            "업무추진을 위한 각종 회의·간담회·행사·교육",
            "식사",
            "의장, 의원",
            4,
            54000,
            "속초명가",
            "13:28",
            "카드",
        ]
    )
    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="서대문구의회")

    assert len(rows) == 1
    assert rows[0].place_text == "속초명가"
    assert rows[0].purpose == "업무추진을 위한 각종 회의·간담회·행사·교육"
    assert rows[0].amount == 54000
    assert rows[0].payment_method == "카드"
    assert rows[0].expense_category == "의회운영"


def test_extracts_xlsx_with_combined_datetime_and_party_size_aliases() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["연번", "사용자", "일시", "장소", "집행목적", "대상인원수", "금액", "결제", "비목"])
    worksheet.append(
        [
            1,
            "체육정책팀",
            "2026-04-01 12:13:00",
            "추오정남원추어탕",
            "양천마라톤 대회 관련 관계자 간담회 비용 지급",
            6,
            54000,
            "신용카드",
            "시책",
        ]
    )

    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="양천구청")

    assert len(rows) == 1
    assert rows[0].place_text == "추오정남원추어탕"
    assert rows[0].used_at.isoformat() == "2026-04-01T12:13:00"
    assert rows[0].amount == 54000
    assert rows[0].user_text == "체육정책팀 6명"


def test_extracts_xlsx_short_year_datetime_as_2000s_year() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["연번", "사용일시", "장소", "집행목적", "금액", "결제방법"])
    worksheet.append([1, "26.04.30, 13:38", "동트팔팔장어", "간담회", 120000, "카드"])

    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="노원구청")

    assert len(rows) == 1
    assert rows[0].used_at.isoformat() == "2026-04-30T13:38:00"
    assert rows[0].place_text == "동트팔팔장어"


def test_extracts_xlsx_short_year_datetime_first_day_without_dateutil_flip() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["연번", "사용일시", "장소", "집행목적", "금액", "결제방법"])
    worksheet.append([1, "26.04.01, 17:00", "노원어르신행복 주식", "간담회", 90000, "카드"])

    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="노원구청")

    assert len(rows) == 1
    assert rows[0].used_at.isoformat() == "2026-04-01T17:00:00"
    assert rows[0].place_text == "노원어르신행복 주식"


def test_extracts_council_cost_xlsx_approval_date_headers_without_place() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "의원회의"
    worksheet.append(["의장단 집행내역"])
    worksheet.append(["구분", "승인일", "승인시각", "승인금액", "대상인원", "집행내역", "결제방법"])
    worksheet.append(
        [
            "의장",
            "2026-01-02 00:00:00",
            "1970-01-01 11:25:19",
            110000,
            5,
            "더불어민주당 의원 간담회",
            "카드결제",
        ]
    )

    content = BytesIO()
    workbook.save(content)

    rows = extract_spreadsheet_rows(content.getvalue(), fallback_department="노원구의회 의장단")

    assert len(rows) == 1
    assert rows[0].place_text == "더불어민주당 의원 간담회"
    assert rows[0].user_text == "의장 5명"
    assert rows[0].amount == 110000


def test_extracts_html_table_served_as_spreadsheet_bytes() -> None:
    content = """
    <html>
      <body>
        <table>
          <tr>
            <th>집행일시</th>
            <th>사용장소</th>
            <th>사용목적</th>
            <th>사용금액(원)</th>
            <th>대상인원</th>
            <th>결제방법</th>
          </tr>
          <tr>
            <td>2026-04-03 12:30</td>
            <td>양천식당</td>
            <td>지역 현안 간담회</td>
            <td>45,000</td>
            <td>3</td>
            <td>카드</td>
          </tr>
        </table>
      </body>
    </html>
    """.encode("cp949")

    rows = extract_spreadsheet_rows(content, fallback_department="양천구청")

    assert len(rows) == 1
    assert rows[0].department_name == "양천구청"
    assert rows[0].used_at.isoformat() == "2026-04-03T12:30:00"
    assert rows[0].place_text == "양천식당"
    assert rows[0].purpose == "지역 현안 간담회"
    assert rows[0].amount == 45000
    assert rows[0].user_text == "양천구청 3명"
    assert rows[0].payment_method == "카드"


def test_extracts_xlsx_with_missing_builtin_cell_style_name() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["집행일자", "장소", "금액"])
    worksheet.append(["2026-04-01", "남양주식당", 33000])

    rows = extract_spreadsheet_rows(
        _workbook_bytes_without_normal_style_name(workbook),
        fallback_department="남양주시의회",
    )

    assert len(rows) == 1
    assert rows[0].department_name == "남양주시의회"
    assert rows[0].place_text == "남양주식당"
    assert rows[0].amount == 33000


def test_rejects_spreadsheet_content_over_limit(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(guards, "MAX_SPREADSHEET_BYTES", 3)

    with pytest.raises(guards.DocumentProcessingLimitError, match="spreadsheet document"):
        extract_spreadsheet_rows(b"not-xlsx", fallback_department="강남구청")


def test_rejects_xlsx_zip_entry_over_limit_before_openpyxl(monkeypatch: pytest.MonkeyPatch) -> None:
    called = False
    monkeypatch.setattr(guards, "MAX_XLSX_ZIP_ENTRY_BYTES", 3)

    def fake_load_workbook(*_args: object, **_kwargs: object) -> object:
        nonlocal called
        called = True
        raise AssertionError("openpyxl should not run after ZIP preflight fails")

    monkeypatch.setattr(spreadsheet_module, "load_workbook", fake_load_workbook)

    with pytest.raises(guards.DocumentProcessingLimitError, match="XLSX ZIP entry"):
        extract_spreadsheet_rows(
            _zip_bytes({"xl/sharedStrings.xml": b"1234"}),
            fallback_department="강남구청",
        )

    assert not called


def test_rejects_xlsx_zip_uncompressed_total_over_limit(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(guards, "MAX_XLSX_ZIP_UNCOMPRESSED_BYTES", 5)

    with pytest.raises(guards.DocumentProcessingLimitError, match="uncompressed total"):
        extract_spreadsheet_rows(
            _zip_bytes({"xl/worksheets/sheet1.xml": b"123", "xl/sharedStrings.xml": b"123"}),
            fallback_department="강남구청",
        )


def test_rejects_xlsx_zip_entry_count_over_limit(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(guards, "MAX_XLSX_ZIP_ENTRIES", 1)

    with pytest.raises(guards.DocumentProcessingLimitError, match="entries"):
        extract_spreadsheet_rows(
            _zip_bytes({"[Content_Types].xml": b"", "xl/workbook.xml": b""}),
            fallback_department="강남구청",
        )


def test_rejects_xlsx_with_too_many_rows(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(guards, "MAX_SPREADSHEET_ROWS_PER_SHEET", 2)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["집행일자", "장소", "금액"])
    worksheet.append(["2026-04-01", "식당", 10000])
    worksheet.append(["2026-04-02", "식당", 10000])

    with pytest.raises(guards.DocumentProcessingLimitError, match="rows"):
        extract_spreadsheet_rows(_workbook_bytes(workbook), fallback_department="강남구청")


def test_extracts_xlsx_with_large_blank_formatted_range() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["집행일자", "장소", "금액"])
    worksheet.append(["2026-04-01", "식당", 10000])
    worksheet.cell(row=guards.MAX_SPREADSHEET_ROWS_PER_SHEET + 1000, column=1).number_format = "@"

    rows = extract_spreadsheet_rows(_workbook_bytes(workbook), fallback_department="서구의회")

    assert len(rows) == 1
    assert rows[0].place_text == "식당"
    assert rows[0].amount == 10000


def test_extracts_xlsx_with_large_blank_formatted_columns() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["집행일자", "장소", "금액"])
    worksheet.append(["2026-04-01", "식당", 10000])
    worksheet.cell(row=1, column=guards.MAX_SPREADSHEET_COLUMNS_PER_SHEET + 1000).number_format = "@"

    rows = extract_spreadsheet_rows(_workbook_bytes(workbook), fallback_department="구리시청")

    assert len(rows) == 1
    assert rows[0].place_text == "식당"
    assert rows[0].amount == 10000


def test_rejects_xlsx_with_too_many_columns(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(guards, "MAX_SPREADSHEET_COLUMNS_PER_SHEET", 2)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["집행일자", "장소", "금액"])

    with pytest.raises(guards.DocumentProcessingLimitError, match="columns"):
        extract_spreadsheet_rows(_workbook_bytes(workbook), fallback_department="강남구청")


def test_extracts_first_xlsx_sheets_within_sheet_budget(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(guards, "MAX_SPREADSHEET_SHEETS", 1)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["집행일자", "장소", "금액"])
    worksheet.append(["2026-04-01", "첫번째식당", 10000])
    second = workbook.create_sheet("second")
    second.append(["집행일자", "장소", "금액"])
    second.append(["2026-04-02", "두번째식당", 20000])

    rows = extract_spreadsheet_rows(_workbook_bytes(workbook), fallback_department="강남구청")

    assert len(rows) == 1
    assert rows[0].place_text == "첫번째식당"


def test_rejects_xlsx_cell_budget(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(guards, "MAX_SPREADSHEET_CELLS_TOTAL", 4)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["집행일자", "장소", "금액"])
    worksheet.append(["2026-04-01", "식당", 10000])

    with pytest.raises(guards.DocumentProcessingLimitError, match="cells"):
        extract_spreadsheet_rows(_workbook_bytes(workbook), fallback_department="강남구청")


def test_rejects_xls_declared_column_budget(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(guards, "MAX_SPREADSHEET_COLUMNS_PER_SHEET", 2)

    class FakeSheet:
        name = "legacy"
        nrows = 1
        ncols = 3

    class FakeWorkbook:
        nsheets = 1
        datemode = 0

        def sheets(self) -> list[FakeSheet]:
            return [FakeSheet()]

    monkeypatch.setattr(
        spreadsheet_module.xlrd,
        "open_workbook",
        lambda *, file_contents: FakeWorkbook(),
    )

    with pytest.raises(guards.DocumentProcessingLimitError, match="columns"):
        extract_spreadsheet_rows(
            b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1legacy",
            fallback_department="강남구청",
        )
